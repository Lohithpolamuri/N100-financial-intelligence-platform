from pathlib import Path
import sqlite3
import pandas as pd


from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent.parent

DB_PATH = BASE_DIR / "db" / "n100.db"

OUTPUT = BASE_DIR / "output"
OUTPUT.mkdir(exist_ok=True)

OUTPUT_FILE = OUTPUT / "peer_comparison.xlsx"
class PeerComparisonReport:

    def __init__(self):



        self.conn = sqlite3.connect(DB_PATH)

        self.financial = pd.read_sql(
            "SELECT * FROM financial_ratios",
            self.conn
        )
        self.cagr = pd.read_sql(
            "SELECT * FROM cagr_metrics",
            self.conn
        )

        self.peer_groups = pd.read_sql(
            "SELECT * FROM peer_groups",
            self.conn
        )

        self.percentiles = pd.read_sql(
            "SELECT * FROM peer_percentiles",
            self.conn
        )



        self.market = pd.read_sql("""
        SELECT company_id,
               company_name
        FROM companies
        """, self.conn)

    def prepare_data(self):
        # Start from financial ratios (contains company_id + year)
        df = self.financial.copy()

            # Add peer group information
        df = df.merge(
            self.cagr,
            on="company_id",
            how="left"
        )

        df=df.merge(
            self.peer_groups,
             on="company_id",
            how="left"
        )


            # Add company names if needed
        if "company_name" not in df.columns:
                df = df.merge(
                    self.market,
                    on="company_id",
                    how="left"
                )

            # Merge percentile table here...



        # Pivot percentile table
        metric_columns = [
            c for c in df.columns
            if c not in [
                "company_id",
                "company_name",
                "year",
                "peer_group_name",
                "is_benchmark"
            ]
        ]

        for metric in metric_columns:

            if pd.api.types.is_numeric_dtype(df[metric]):

                higher_is_better = metric != "debt_to_equity"

                if higher_is_better:
                    df[f"{metric}_percentile"] = (
                            df.groupby("peer_group_name")[metric]
                            .rank(pct=True) * 100
                    )
                else:
                    df[f"{metric}_percentile"] = (
                            (1 - df.groupby("peer_group_name")[metric].rank(pct=True))
                            * 100
                    )

        self.report_df = df

    def generate_excel(self):

        wb = Workbook()
        wb.remove(wb.active)

        green = PatternFill(fill_type="solid", fgColor="90EE90")
        yellow = PatternFill(fill_type="solid", fgColor="FFF59D")
        red = PatternFill(fill_type="solid", fgColor="FF9999")
        gold = PatternFill(fill_type="solid", fgColor="FFD966")

        percentile_cols = [c for c in self.report_df.columns if c.endswith("_percentile")]

        metric_cols = [
            c for c in self.report_df.columns
            if c not in percentile_cols
               and c not in ["peer_group_name", "is_benchmark"]
        ]

        for peer_group in sorted(self.report_df["peer_group_name"].dropna().unique()):

            ws = wb.create_sheet(title=str(peer_group)[:31])

            df = self.report_df[self.report_df["peer_group_name"] == peer_group].copy()

            columns = metric_cols + percentile_cols

            ws.append(columns)

            for row in df[columns].itertuples(index=False):
                ws.append(list(row))

            # Highlight benchmark row
            for r in range(2, ws.max_row + 1):

                company = ws.cell(r, 1).value

                benchmark = df.loc[
                    df["company_id"] == company,
                    "is_benchmark"
                ].iloc[0]

                if benchmark == 1:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).fill = gold

            # Percentile colors
            for col in percentile_cols:

                col_index = columns.index(col) + 1

                for r in range(2, ws.max_row + 1):

                    value = ws.cell(r, col_index).value

                    if value is None:
                        continue

                    if value >= 75:
                        ws.cell(r, col_index).fill = green

                    elif value <= 25:
                        ws.cell(r, col_index).fill = red

                    else:
                        ws.cell(r, col_index).fill = yellow

            # Median row
            median_row = ["Median"]

            for col in columns[1:]:

                if pd.api.types.is_numeric_dtype(df[col]):
                    median_row.append(round(df[col].median(), 2))
                else:
                    median_row.append("")

            ws.append(median_row)

            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)

            # Auto width
            for column in ws.columns:
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
                ws.column_dimensions[get_column_letter(column[0].column)].width = min(length + 2, 30)

        wb.save(OUTPUT_FILE)

        print(f"Report saved to: {OUTPUT_FILE}")
if __name__ == "__main__":

    report = PeerComparisonReport()

    report.prepare_data()


    report.generate_excel()