from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


from pathlib import Path
import pandas as pd

from src.screener.engine import ScreenerEngine
from src.screener.presets import PresetScreeners

# ==========================================================
# Paths
# ==========================================================

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "output"

OUTPUT_DIR.mkdir(exist_ok=True)

OUTPUT_FILE = OUTPUT_DIR / "screener_output.xlsx"

# ==========================================================
# Load Engine
# ==========================================================

engine = ScreenerEngine()

presets = PresetScreeners(engine.master_df)

# ==========================================================
# Run All Screeners
# ==========================================================

screeners = {
    "Quality Compounder": presets.quality_compounder(),
    "Value Pick": presets.value_pick(),
    "Growth Accelerator": presets.growth_accelerator(),
    "Dividend Champion": presets.dividend_champion(),
    "Debt Free Blue Chip": presets.debt_free_bluechip(),
    "Turnaround Watch": presets.turnaround_watch()
}

# ==========================================================
# Export
# ==========================================================

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

    for sheet_name, df in screeners.items():

        df = df.sort_values(
            "composite_quality_score",
            ascending=False
        )

        columns = [
            "company_id",
            "year",
            "composite_quality_score",
            "return_on_equity_pct",
            "net_profit_margin_pct",
            "operating_profit_margin_pct",
            "debt_to_equity",
            "interest_coverage",
            "free_cash_flow_cr",
            "sales",
            "net_profit",
            "market_cap_crore",
            "pe_ratio",
            "pb_ratio",
            "dividend_yield_pct",
            "revenue_cagr_5yr",
            "pat_cagr_5yr",
            "eps_cagr_5yr",
            "broad_sector"
        ]

        available_columns = [
            c for c in columns
            if c in df.columns
        ]

        df[available_columns].to_excel(
            writer,
            sheet_name=sheet_name,
            index=False
        )
        worksheet = writer.sheets[sheet_name]

        # -----------------------------
        # Header Formatting
        # -----------------------------
        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        # -----------------------------
        # KPI Cell Colours
        # -----------------------------
        green_fill = PatternFill(
            fill_type="solid",
            start_color="C6EFCE",
            end_color="C6EFCE"
        )

        red_fill = PatternFill(
            fill_type="solid",
            start_color="FFC7CE",
            end_color="FFC7CE"
        )
        # -----------------------------
        # Colour KPI Cells
        # -----------------------------

        thresholds = {
            "return_on_equity_pct": 15,
            "debt_to_equity": 1,
            "free_cash_flow_cr": 0,
            "revenue_cagr_5yr": 10,
            "pat_cagr_5yr": 10,
            "operating_profit_margin_pct": 10,
            "pe_ratio": 20,
            "pb_ratio": 3,
            "dividend_yield_pct": 1,
            "interest_coverage": 2,
            "market_cap_crore": 5000,
            "net_profit": 100,
            "eps_cagr_5yr": 10,
            "sales": 1000
        }

        reverse_metrics = {
            "debt_to_equity",
            "pe_ratio",
            "pb_ratio"
        }

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        for metric, threshold in thresholds.items():

            if metric not in headers:
                continue

            col = headers[metric]

            for row in range(2, worksheet.max_row + 1):

                cell = worksheet.cell(row=row, column=col)

                if cell.value is None:
                    continue

                try:
                    value = float(cell.value)
                except:
                    continue

                if metric in reverse_metrics:
                    cell.fill = (
                        green_fill
                        if value <= threshold
                        else red_fill
                    )
                else:
                    cell.fill = (
                        green_fill
                        if value >= threshold
                        else red_fill
                    )

        # -----------------------------
        # Freeze Header Row
        # -----------------------------
        worksheet.freeze_panes = "A2"

        # -----------------------------
        # Auto-fit Columns
        # -----------------------------
        for column_cells in worksheet.columns:
            length = max(
                len(str(cell.value))
                if cell.value is not None else 0
                for cell in column_cells
            )

            worksheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = length + 3

print("=" * 60)
print("SCREENERS EXPORTED SUCCESSFULLY")
print("=" * 60)
print(OUTPUT_FILE)